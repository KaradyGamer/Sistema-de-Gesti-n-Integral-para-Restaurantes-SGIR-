# ═══════════════════════════════════════════
# RONDA 4: Exportación de Reportes PDF/Excel
# ═══════════════════════════════════════════

from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from django.http import HttpResponse
from django.db.models import Sum, Count, F
from django.utils import timezone
from decimal import Decimal
import logging

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet

from app.pedidos.models import Pedido, DetallePedido
from app.caja.models import CierreCaja, Reembolso
from .utils import parse_rango_fechas, require_admin_or_manager, qp_bool, qp_int, qp_choice

logger = logging.getLogger('app.reportes')


# ═══════════════════════════════════════════
# REPORTE DE VENTAS
# ═══════════════════════════════════════════

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def ventas_xlsx(request):
    """Exporta reporte de ventas a Excel (.xlsx)"""
    # Validar permisos
    error_response = require_admin_or_manager(request.user)
    if error_response:
        return error_response

    # Parsear rango de fechas
    fecha_inicio, fecha_fin = parse_rango_fechas(request)
    if not fecha_inicio or not fecha_fin:
        return Response({'error': 'Rango de fechas inválido'}, status=400)

    # RONDA 4.1: Filtro opcional solo_cerrados
    solo_cerrados = qp_bool(request, "solo_cerrados", default=False)

    # Consultar pedidos en el rango
    pedidos = Pedido.objects.filter(
        fecha_pago__date__gte=fecha_inicio.date(),
        fecha_pago__date__lte=fecha_fin.date()
    ).select_related('mesa', 'cajero_responsable', 'mesero_comanda')

    if solo_cerrados:
        pedidos = pedidos.filter(estado='cerrado')
    else:
        # Por defecto solo pedidos cerrados (comportamiento original)
        pedidos = pedidos.filter(estado='cerrado')

    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte de Ventas"

    # Estilos
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Título
    ws.merge_cells('A1:J1')
    ws['A1'] = f"Reporte de Ventas - {fecha_inicio.strftime('%d/%m/%Y')} a {fecha_fin.strftime('%d/%m/%Y')}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')

    # Encabezados
    headers = ['ID Pedido', 'Mesa', 'Mesero', 'Cajero', 'Fecha Pago', 'Total', 'Descuento', 'Propina', 'Total Final', 'Forma Pago']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center')

    # Datos
    row = 4
    total_ventas = Decimal('0.00')
    total_descuentos = Decimal('0.00')
    total_propinas = Decimal('0.00')
    total_final_sum = Decimal('0.00')

    for pedido in pedidos:
        ws.cell(row=row, column=1, value=pedido.id).border = border
        ws.cell(row=row, column=2, value=pedido.mesa.numero if pedido.mesa else 'N/A').border = border
        ws.cell(row=row, column=3, value=pedido.mesero_comanda.username if pedido.mesero_comanda else 'N/A').border = border
        ws.cell(row=row, column=4, value=pedido.cajero_responsable.username if pedido.cajero_responsable else 'N/A').border = border
        ws.cell(row=row, column=5, value=pedido.fecha_pago.strftime('%d/%m/%Y %H:%M') if pedido.fecha_pago else 'N/A').border = border
        ws.cell(row=row, column=6, value=float(pedido.total)).border = border
        ws.cell(row=row, column=7, value=float(pedido.descuento)).border = border
        ws.cell(row=row, column=8, value=float(pedido.propina)).border = border
        ws.cell(row=row, column=9, value=float(pedido.total_final)).border = border
        ws.cell(row=row, column=10, value=pedido.forma_pago or 'N/A').border = border

        total_ventas += pedido.total
        total_descuentos += pedido.descuento
        total_propinas += pedido.propina
        total_final_sum += pedido.total_final
        row += 1

    # Totales
    row += 1
    ws.cell(row=row, column=5, value='TOTALES:').font = Font(bold=True)
    ws.cell(row=row, column=6, value=float(total_ventas)).font = Font(bold=True)
    ws.cell(row=row, column=7, value=float(total_descuentos)).font = Font(bold=True)
    ws.cell(row=row, column=8, value=float(total_propinas)).font = Font(bold=True)
    ws.cell(row=row, column=9, value=float(total_final_sum)).font = Font(bold=True)

    # Ajustar anchos
    for col in range(1, 11):
        ws.column_dimensions[chr(64 + col)].width = 15

    # Generar respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="ventas_{fecha_inicio.strftime("%Y%m%d")}_{fecha_fin.strftime("%Y%m%d")}.xlsx"'
    wb.save(response)

    logger.info(f"AUDIT reporte_ventas_xlsx user={request.user.username} desde={fecha_inicio.date()} hasta={fecha_fin.date()} pedidos={pedidos.count()}")

    return response


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def ventas_pdf(request):
    """Exporta reporte de ventas a PDF"""
    # Validar permisos
    error_response = require_admin_or_manager(request.user)
    if error_response:
        return error_response

    # Parsear rango de fechas
    fecha_inicio, fecha_fin = parse_rango_fechas(request)
    if not fecha_inicio or not fecha_fin:
        return Response({'error': 'Rango de fechas inválido'}, status=400)

    # RONDA 4.1: Filtro opcional solo_cerrados
    solo_cerrados = qp_bool(request, "solo_cerrados", default=False)

    # Consultar pedidos en el rango
    pedidos = Pedido.objects.filter(
        fecha_pago__date__gte=fecha_inicio.date(),
        fecha_pago__date__lte=fecha_fin.date()
    ).select_related('mesa', 'cajero_responsable', 'mesero_comanda')

    if solo_cerrados:
        pedidos = pedidos.filter(estado='cerrado')
    else:
        # Por defecto solo pedidos cerrados (comportamiento original)
        pedidos = pedidos.filter(estado='cerrado')

    # Crear PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="ventas_{fecha_inicio.strftime("%Y%m%d")}_{fecha_fin.strftime("%Y%m%d")}.pdf"'

    doc = SimpleDocTemplate(response, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()

    # Título
    titulo = Paragraph(
        f"<b>Reporte de Ventas</b><br/>{fecha_inicio.strftime('%d/%m/%Y')} - {fecha_fin.strftime('%d/%m/%Y')}",
        styles['Title']
    )
    elements.append(titulo)
    elements.append(Spacer(1, 0.3 * inch))

    # Tabla
    data = [['ID', 'Mesa', 'Mesero', 'Cajero', 'Fecha', 'Total', 'Descuento', 'Propina', 'Total Final']]

    total_ventas = Decimal('0.00')
    total_descuentos = Decimal('0.00')
    total_propinas = Decimal('0.00')
    total_final_sum = Decimal('0.00')

    for pedido in pedidos:
        data.append([
            str(pedido.id),
            pedido.mesa.numero if pedido.mesa else 'N/A',
            pedido.mesero_comanda.username if pedido.mesero_comanda else 'N/A',
            pedido.cajero_responsable.username if pedido.cajero_responsable else 'N/A',
            pedido.fecha_pago.strftime('%d/%m/%Y') if pedido.fecha_pago else 'N/A',
            f"${float(pedido.total):.2f}",
            f"${float(pedido.descuento):.2f}",
            f"${float(pedido.propina):.2f}",
            f"${float(pedido.total_final):.2f}",
        ])
        total_ventas += pedido.total
        total_descuentos += pedido.descuento
        total_propinas += pedido.propina
        total_final_sum += pedido.total_final

    # Fila de totales
    data.append([
        '', '', '', '', 'TOTALES:',
        f"${float(total_ventas):.2f}",
        f"${float(total_descuentos):.2f}",
        f"${float(total_propinas):.2f}",
        f"${float(total_final_sum):.2f}"
    ])

    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F81BD')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#D9E1F2')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))

    elements.append(table)
    doc.build(elements)

    logger.info(f"AUDIT reporte_ventas_pdf user={request.user.username} desde={fecha_inicio.date()} hasta={fecha_fin.date()} pedidos={pedidos.count()}")

    return response


# ═══════════════════════════════════════════
# REPORTE DE CAJA (TURNOS)
# ═══════════════════════════════════════════

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def caja_xlsx(request):
    """Exporta reporte de turnos de caja a Excel"""
    error_response = require_admin_or_manager(request.user)
    if error_response:
        return error_response

    fecha_inicio, fecha_fin = parse_rango_fechas(request)
    if not fecha_inicio or not fecha_fin:
        return Response({'error': 'Rango de fechas inválido'}, status=400)

    # RONDA 4 HOTFIX: usar campos reales de CierreCaja
    turnos = CierreCaja.objects.filter(
        fecha__range=(fecha_inicio.date(), fecha_fin.date())
    ).select_related('cajero')

    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte de Caja"

    # Estilos
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Título
    ws.merge_cells('A1:I1')
    ws['A1'] = f"Reporte de Turnos de Caja - {fecha_inicio.strftime('%d/%m/%Y')} a {fecha_fin.strftime('%d/%m/%Y')}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')

    # Encabezados
    headers = ['ID', 'Cajero', 'Fecha', 'Turno', 'Efectivo Inicial', 'Total Efectivo', 'Total Tarjeta', 'Total QR', 'Total Ventas']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center')

    # Datos
    row = 4
    total_acumulado = Decimal('0.00')

    for turno in turnos:
        # Usar campos reales
        total_ventas = getattr(turno, 'total_ventas', 0) or 0

        ws.cell(row=row, column=1, value=turno.id).border = border
        ws.cell(row=row, column=2, value=turno.cajero.username if turno.cajero else 'N/A').border = border
        ws.cell(row=row, column=3, value=turno.fecha.strftime('%d/%m/%Y')).border = border
        ws.cell(row=row, column=4, value=turno.get_turno_display()).border = border
        ws.cell(row=row, column=5, value=float(turno.efectivo_inicial or 0)).border = border
        ws.cell(row=row, column=6, value=float(turno.total_efectivo or 0)).border = border
        ws.cell(row=row, column=7, value=float(turno.total_tarjeta or 0)).border = border
        ws.cell(row=row, column=8, value=float(turno.total_qr or 0)).border = border
        ws.cell(row=row, column=9, value=float(total_ventas)).border = border

        total_acumulado += Decimal(str(total_ventas))
        row += 1

    # Totales
    row += 1
    ws.cell(row=row, column=8, value='TOTAL:').font = Font(bold=True)
    ws.cell(row=row, column=9, value=float(total_acumulado)).font = Font(bold=True)

    for col in range(1, 9):
        ws.column_dimensions[chr(64 + col)].width = 15

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="caja_{fecha_inicio.strftime("%Y%m%d")}_{fecha_fin.strftime("%Y%m%d")}.xlsx"'
    wb.save(response)

    logger.info(f"AUDIT reporte_caja_xlsx user={request.user.username} desde={fecha_inicio.date()} hasta={fecha_fin.date()} turnos={turnos.count()}")

    return response


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def caja_pdf(request):
    """Exporta reporte de turnos de caja a PDF"""
    error_response = require_admin_or_manager(request.user)
    if error_response:
        return error_response

    fecha_inicio, fecha_fin = parse_rango_fechas(request)
    if not fecha_inicio or not fecha_fin:
        return Response({'error': 'Rango de fechas inválido'}, status=400)

    # RONDA 4 HOTFIX: usar campos reales de CierreCaja
    turnos = CierreCaja.objects.filter(
        fecha__range=(fecha_inicio.date(), fecha_fin.date())
    ).select_related('cajero')

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="caja_{fecha_inicio.strftime("%Y%m%d")}_{fecha_fin.strftime("%Y%m%d")}.pdf"'

    doc = SimpleDocTemplate(response, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()

    titulo = Paragraph(
        f"<b>Reporte de Turnos de Caja</b><br/>{fecha_inicio.strftime('%d/%m/%Y')} - {fecha_fin.strftime('%d/%m/%Y')}",
        styles['Title']
    )
    elements.append(titulo)
    elements.append(Spacer(1, 0.3 * inch))

    data = [['ID', 'Cajero', 'Fecha', 'Turno', 'Inicial', 'Efectivo', 'Tarjeta', 'QR', 'Total']]

    total_acumulado = Decimal('0.00')

    for turno in turnos:
        # Usar campos reales
        total_ventas = getattr(turno, 'total_ventas', 0) or 0

        data.append([
            str(turno.id),
            turno.cajero.username if turno.cajero else 'N/A',
            turno.fecha.strftime('%d/%m/%Y'),
            turno.get_turno_display(),
            f"${float(turno.efectivo_inicial or 0):.2f}",
            f"${float(turno.total_efectivo or 0):.2f}",
            f"${float(turno.total_tarjeta or 0):.2f}",
            f"${float(turno.total_qr or 0):.2f}",
            f"${float(total_ventas):.2f}",
        ])
        total_acumulado += Decimal(str(total_ventas))

    data.append(['', '', '', '', '', '', '', 'TOTAL:', f"${float(total_acumulado):.2f}"])

    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F81BD')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#D9E1F2')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))

    elements.append(table)
    doc.build(elements)

    logger.info(f"AUDIT reporte_caja_pdf user={request.user.username} desde={fecha_inicio.date()} hasta={fecha_fin.date()} turnos={turnos.count()}")

    return response


# ═══════════════════════════════════════════
# REPORTE DE REEMBOLSOS
# ═══════════════════════════════════════════

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def reembolsos_xlsx(request):
    """Exporta reporte de reembolsos a Excel"""
    error_response = require_admin_or_manager(request.user)
    if error_response:
        return error_response

    fecha_inicio, fecha_fin = parse_rango_fechas(request)
    if not fecha_inicio or not fecha_fin:
        return Response({'error': 'Rango de fechas inválido'}, status=400)

    # RONDA 4.1: Filtro opcional por método
    metodo = qp_choice(request, "metodo", {"efectivo", "qr", "tarjeta", "movil"})

    reembolsos = Reembolso.objects.filter(
        creado_en__date__gte=fecha_inicio.date(),
        creado_en__date__lte=fecha_fin.date()
    ).select_related('pedido', 'creado_por', 'autorizado_por')

    if metodo:
        reembolsos = reembolsos.filter(metodo=metodo)

    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte de Reembolsos"

    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    ws.merge_cells('A1:G1')
    ws['A1'] = f"Reporte de Reembolsos - {fecha_inicio.strftime('%d/%m/%Y')} a {fecha_fin.strftime('%d/%m/%Y')}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')

    headers = ['ID Reembolso', 'ID Pedido', 'Monto', 'Método', 'Motivo', 'Autorizado Por', 'Fecha']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center')

    row = 4
    total_reembolsado = Decimal('0.00')

    for reembolso in reembolsos:
        ws.cell(row=row, column=1, value=reembolso.id).border = border
        ws.cell(row=row, column=2, value=reembolso.pedido.id).border = border
        ws.cell(row=row, column=3, value=float(reembolso.monto)).border = border
        ws.cell(row=row, column=4, value=reembolso.metodo).border = border
        ws.cell(row=row, column=5, value=reembolso.motivo[:50]).border = border
        ws.cell(row=row, column=6, value=reembolso.autorizado_por.username if reembolso.autorizado_por else 'N/A').border = border
        ws.cell(row=row, column=7, value=reembolso.creado_en.strftime('%d/%m/%Y %H:%M')).border = border

        total_reembolsado += reembolso.monto
        row += 1

    row += 1
    ws.cell(row=row, column=2, value='TOTAL REEMBOLSADO:').font = Font(bold=True)
    ws.cell(row=row, column=3, value=float(total_reembolsado)).font = Font(bold=True)

    for col in range(1, 8):
        ws.column_dimensions[chr(64 + col)].width = 15

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="reembolsos_{fecha_inicio.strftime("%Y%m%d")}_{fecha_fin.strftime("%Y%m%d")}.xlsx"'
    wb.save(response)

    logger.info(f"AUDIT reporte_reembolsos_xlsx user={request.user.username} desde={fecha_inicio.date()} hasta={fecha_fin.date()} reembolsos={reembolsos.count()}")

    return response


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def reembolsos_pdf(request):
    """Exporta reporte de reembolsos a PDF"""
    error_response = require_admin_or_manager(request.user)
    if error_response:
        return error_response

    fecha_inicio, fecha_fin = parse_rango_fechas(request)
    if not fecha_inicio or not fecha_fin:
        return Response({'error': 'Rango de fechas inválido'}, status=400)

    # RONDA 4.1: Filtro opcional por método
    metodo = qp_choice(request, "metodo", {"efectivo", "qr", "tarjeta", "movil"})

    reembolsos = Reembolso.objects.filter(
        creado_en__date__gte=fecha_inicio.date(),
        creado_en__date__lte=fecha_fin.date()
    ).select_related('pedido', 'creado_por', 'autorizado_por')

    if metodo:
        reembolsos = reembolsos.filter(metodo=metodo)

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="reembolsos_{fecha_inicio.strftime("%Y%m%d")}_{fecha_fin.strftime("%Y%m%d")}.pdf"'

    doc = SimpleDocTemplate(response, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()

    titulo = Paragraph(
        f"<b>Reporte de Reembolsos</b><br/>{fecha_inicio.strftime('%d/%m/%Y')} - {fecha_fin.strftime('%d/%m/%Y')}",
        styles['Title']
    )
    elements.append(titulo)
    elements.append(Spacer(1, 0.3 * inch))

    data = [['ID', 'Pedido', 'Monto', 'Método', 'Motivo', 'Autorizado Por', 'Fecha']]

    total_reembolsado = Decimal('0.00')

    for reembolso in reembolsos:
        data.append([
            str(reembolso.id),
            str(reembolso.pedido.id),
            f"${float(reembolso.monto):.2f}",
            reembolso.metodo,
            reembolso.motivo[:30],
            reembolso.autorizado_por.username if reembolso.autorizado_por else 'N/A',
            reembolso.creado_en.strftime('%d/%m/%Y'),
        ])
        total_reembolsado += reembolso.monto

    data.append(['', 'TOTAL:', f"${float(total_reembolsado):.2f}", '', '', '', ''])

    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F81BD')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#D9E1F2')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))

    elements.append(table)
    doc.build(elements)

    logger.info(f"AUDIT reporte_reembolsos_pdf user={request.user.username} desde={fecha_inicio.date()} hasta={fecha_fin.date()} reembolsos={reembolsos.count()}")

    return response


# ═══════════════════════════════════════════
# REPORTE DE TOP PRODUCTOS
# ═══════════════════════════════════════════

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def top_productos_xlsx(request):
    """Exporta top productos más vendidos a Excel"""
    error_response = require_admin_or_manager(request.user)
    if error_response:
        return error_response

    fecha_inicio, fecha_fin = parse_rango_fechas(request)
    if not fecha_inicio or not fecha_fin:
        return Response({'error': 'Rango de fechas inválido'}, status=400)

    # RONDA 4.1: Filtro opcional top (cantidad de productos)
    top_n = qp_int(request, "top", default=20, min_v=1, max_v=100)

    # Top N productos más vendidos
    productos = DetallePedido.objects.filter(
        pedido__estado='cerrado',
        pedido__fecha_pago__date__gte=fecha_inicio.date(),
        pedido__fecha_pago__date__lte=fecha_fin.date()
    ).values(
        'producto__nombre'
    ).annotate(
        cantidad_vendida=Sum('cantidad'),
        ingresos_totales=Sum('subtotal')
    ).order_by('-cantidad_vendida')[:top_n]

    wb = Workbook()
    ws = wb.active
    ws.title = "Top Productos"

    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    ws.merge_cells('A1:D1')
    ws['A1'] = f"Top {top_n} Productos Más Vendidos - {fecha_inicio.strftime('%d/%m/%Y')} a {fecha_fin.strftime('%d/%m/%Y')}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')

    headers = ['Ranking', 'Producto', 'Cantidad Vendida', 'Ingresos Totales']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center')

    row = 4
    ranking = 1
    total_cantidad = 0
    total_ingresos = Decimal('0.00')

    for producto in productos:
        ws.cell(row=row, column=1, value=ranking).border = border
        ws.cell(row=row, column=2, value=producto['producto__nombre']).border = border
        ws.cell(row=row, column=3, value=producto['cantidad_vendida']).border = border
        ws.cell(row=row, column=4, value=float(producto['ingresos_totales'])).border = border

        total_cantidad += producto['cantidad_vendida']
        total_ingresos += producto['ingresos_totales']
        ranking += 1
        row += 1

    row += 1
    ws.cell(row=row, column=2, value='TOTALES:').font = Font(bold=True)
    ws.cell(row=row, column=3, value=total_cantidad).font = Font(bold=True)
    ws.cell(row=row, column=4, value=float(total_ingresos)).font = Font(bold=True)

    for col in range(1, 5):
        ws.column_dimensions[chr(64 + col)].width = 20

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="top_productos_{fecha_inicio.strftime("%Y%m%d")}_{fecha_fin.strftime("%Y%m%d")}.xlsx"'
    wb.save(response)

    logger.info(f"AUDIT reporte_top_productos_xlsx user={request.user.username} desde={fecha_inicio.date()} hasta={fecha_fin.date()} productos={productos.count()}")

    return response


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def top_productos_pdf(request):
    """Exporta top productos más vendidos a PDF"""
    error_response = require_admin_or_manager(request.user)
    if error_response:
        return error_response

    fecha_inicio, fecha_fin = parse_rango_fechas(request)
    if not fecha_inicio or not fecha_fin:
        return Response({'error': 'Rango de fechas inválido'}, status=400)

    # RONDA 4.1: Filtro opcional top (cantidad de productos)
    top_n = qp_int(request, "top", default=20, min_v=1, max_v=100)

    productos = DetallePedido.objects.filter(
        pedido__estado='cerrado',
        pedido__fecha_pago__date__gte=fecha_inicio.date(),
        pedido__fecha_pago__date__lte=fecha_fin.date()
    ).values(
        'producto__nombre'
    ).annotate(
        cantidad_vendida=Sum('cantidad'),
        ingresos_totales=Sum('subtotal')
    ).order_by('-cantidad_vendida')[:top_n]

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="top_productos_{fecha_inicio.strftime("%Y%m%d")}_{fecha_fin.strftime("%Y%m%d")}.pdf"'

    doc = SimpleDocTemplate(response, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()

    titulo = Paragraph(
        f"<b>Top {top_n} Productos Más Vendidos</b><br/>{fecha_inicio.strftime('%d/%m/%Y')} - {fecha_fin.strftime('%d/%m/%Y')}",
        styles['Title']
    )
    elements.append(titulo)
    elements.append(Spacer(1, 0.3 * inch))

    data = [['Ranking', 'Producto', 'Cantidad', 'Ingresos']]

    ranking = 1
    total_cantidad = 0
    total_ingresos = Decimal('0.00')

    for producto in productos:
        data.append([
            str(ranking),
            producto['producto__nombre'],
            str(producto['cantidad_vendida']),
            f"${float(producto['ingresos_totales']):.2f}",
        ])
        total_cantidad += producto['cantidad_vendida']
        total_ingresos += producto['ingresos_totales']
        ranking += 1

    data.append(['', 'TOTALES:', str(total_cantidad), f"${float(total_ingresos):.2f}"])

    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F81BD')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#D9E1F2')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))

    elements.append(table)
    doc.build(elements)

    logger.info(f"AUDIT reporte_top_productos_pdf user={request.user.username} desde={fecha_inicio.date()} hasta={fecha_fin.date()} productos={productos.count()}")

    return response